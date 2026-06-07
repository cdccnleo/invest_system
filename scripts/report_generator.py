"""
report_generator.py — 报告自动生成模块
支持生成 Markdown/HTML/PDF 格式的周报和月报
基于持仓数据、因子评分、新闻摘要自动组装
"""

import logging
from datetime import date, timedelta
from pathlib import Path

logger = logging.getLogger("invest_system.report_generator")

REPORT_DIR = Path(__file__).parent.parent / "reports" / "auto"

REPORT_TEMPLATE = """# {title}

> 生成时间: {generated_at} | 数据截止: {cutoff_date}

---

## 一、市场概览

{market_overview}

## 二、持仓分析

{portfolio_section}

## 三、因子评分排名

{factor_section}

## 四、近期新闻摘要

{news_section}

## 五、操作计划

{plan_section}

## 六、风险提示

{risk_section}

---

*本报告由 InvestPilot 自动生成*
"""


def _fmt_money(value: float) -> str:
    """金额格式化"""
    if abs(value) >= 1e8:
        return f"{value / 1e8:.2f}亿"
    elif abs(value) >= 1e4:
        return f"{value / 1e4:.2f}万"
    return f"{value:.2f}"


def generate_weekly_report() -> str:
    """
    生成投资周报 (Markdown)

    Returns:
        Markdown 格式的周报内容
    """
    today = date.today()
    cutoff = today - timedelta(days=7)

    # 持仓数据
    portfolio_rows = []
    try:
        from dashboard_views._shared import load_positions
        positions = load_positions()
        for p in (positions or []):
            portfolio_rows.append(
                f"| {p.get('code', '—')} | {p.get('name', '—')} | "
                f"{p.get('close', 0):.2f} | {p.get('change_pct', 0):+.2f}% | "
                f"{p.get('market_value', 0):.0f} |"
            )
    except Exception:
        portfolio_rows = ["| — | 暂无持仓数据 | — | — | — |"]

    portfolio_section = (
        "| 代码 | 名称 | 最新价 | 涨跌幅 | 市值 |\n"
        "|------|------|--------|--------|------|\n"
        + "\n".join(portfolio_rows)
    )

    # 因子评分
    factor_section = "暂无因子评分数据"
    try:
        from factor_engine import score_positions
        positions = (lambda: None)()
        try:
            from dashboard_views._shared import load_positions
            positions = load_positions()
        except Exception:
            pass
        if positions:
            results = score_positions(positions)
            if results:
                factor_rows = [
                    f"| {r['rank']} | {r['ts_code']} | {r['total_score']:.1f} | {r['z_score']:+.1f} |"  # noqa: E501
                    for r in results[:10]
                ]
                factor_section = (
                    "| 排名 | 代码 | 综合得分 | Z-Score |\n"
                    "|:---:|------|:---:|:---:|\n"
                    + "\n".join(factor_rows)
                )
    except Exception:
        pass

    # 新闻摘要
    news_section = "暂无近期新闻"
    try:
        from report_summarizer import get_reports_with_summary
        reports = get_reports_with_summary(days=7, limit=10)
        if reports:
            news_lines = []
            for r in reports:
                news_lines.append(f"- **{r.get('title', '—')}** ({r.get('org_name', '—')})")
                if r.get("summary"):
                    news_lines.append(f"  > {r['summary']}")
            news_section = "\n".join(news_lines[:20])
    except Exception:
        pass

    # 组装报告
    report = REPORT_TEMPLATE.format(
        title=f"InvestPilot 投资周报 ({cutoff} ~ {today})",
        generated_at=today.isoformat(),
        cutoff_date=cutoff.isoformat(),
        market_overview="本周市场概述（待数据补充）",
        portfolio_section=portfolio_section,
        factor_section=factor_section,
        news_section=news_section,
        plan_section="待分析完成后更新",
        risk_section="待风险评估后更新",
    )

    return report


def save_report_as_md(report: str, report_type: str = "weekly") -> Path:
    """
    保存报告为 Markdown 文件

    Args:
        report: 报告内容
        report_type: 报告类型 (weekly/monthly)

    Returns:
        保存的文件路径
    """
    REPORT_DIR.mkdir(parents=True, exist_ok=True)
    today = date.today().strftime("%Y%m%d")
    filename = f"{report_type}_{today}.md"
    filepath = REPORT_DIR / filename
    filepath.write_text(report, encoding="utf-8")
    logger.info(f"报告已保存: {filepath}")
    return filepath


def generate_monthly_report() -> str:
    """
    生成投资月报 (Markdown)

    Returns:
        Markdown 格式的月报内容
    """
    today = date.today()
    cutoff = today.replace(day=1) - timedelta(days=1)
    cutoff_start = cutoff.replace(day=1)

    report = REPORT_TEMPLATE.format(
        title=f"InvestPilot 投资月报 ({cutoff_start} ~ {cutoff})",
        generated_at=today.isoformat(),
        cutoff_date=cutoff.isoformat(),
        market_overview="本月市场概述（待数据补充）",
        portfolio_section="本月持仓变化（待数据补充）",
        factor_section="本月因子评分（待数据补充）",
        news_section="本月重要新闻（待数据补充）",
        plan_section="下月操作计划（待分析完成后更新）",
        risk_section="月度风险评估（待评估后更新）",
    )

    return report


def get_recent_reports(report_type: str = "weekly", limit: int = 5) -> list[dict]:
    """
    获取最近的报告列表

    Args:
        report_type: 报告类型
        limit: 返回数量

    Returns:
        报告文件信息列表
    """
    if not REPORT_DIR.exists():
        return []

    files = sorted(
        REPORT_DIR.glob(f"{report_type}_*.md"),
        reverse=True,
    )
    return [
        {
            "name": f.name,
            "path": str(f),
            "size_kb": round(f.stat().st_size / 1024, 1),
            "modified": date.fromtimestamp(f.stat().st_mtime).isoformat(),
        }
        for f in files[:limit]
    ]


def generate_excel_report(positions: list[dict], output_path: str = None) -> str:
    """
    生成Excel报告（持仓/盈亏/因子三sheet）

    Args:
        positions: 持仓数据列表
        output_path: 输出文件路径（可选）

    Returns:
        保存的文件路径
    """
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    wb = openpyxl.Workbook()

    # ── Sheet 1: 持仓明细 ────────────────────────────────────────────────────
    ws1 = wb.active
    ws1.title = "持仓明细"

    headers1 = ["代码", "名称", "份额", "成本", "市值", "盈亏", "盈亏%", "仓位%"]
    header_fill = PatternFill("solid", fgColor="4472C4")
    header_font = Font(bold=True, color="FFFFFF")
    thin = Side(border_style="thin", color="AAAAAA")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    for col, header in enumerate(headers1, start=1):
        cell = ws1.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = border

    # 数据行
    for row_idx, p in enumerate(positions, start=2):
        code = p.get("code", p.get("代码", "—"))
        name = p.get("name", p.get("名称", "—"))
        shares = p.get("shares", p.get("份额", 0))
        cost = p.get("avg_cost", p.get("成本", 0))
        mv = p.get("market_value", p.get("市值", 0))
        profit = mv - shares * cost
        profit_pct = (profit / (shares * cost) * 100) if shares * cost > 0 else 0
        weight = p.get("weight", p.get("仓位%", 0))

        row_data = [code, name, shares, cost, mv, profit, profit_pct, weight]
        for col, value in enumerate(row_data, start=1):
            cell = ws1.cell(row=row_idx, column=col, value=value)
            cell.border = border
            if col >= 3:
                cell.number_format = "#,##0.00"

    # 列宽自适应
    for col in ws1.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            try:
                if cell.value:
                    max_len = max(max_len, len(str(cell.value)))
            except Exception:
                pass
        ws1.column_dimensions[col_letter].width = min(max_len + 4, 20)

    # ── Sheet 2: 盈亏统计 ────────────────────────────────────────────────────
    ws2 = wb.create_sheet("盈亏统计")

    total_cost = sum(p.get("shares", p.get("份额", 0)) * p.get("avg_cost", p.get("成本", 0)) for p in positions)
    total_mv = sum(p.get("market_value", p.get("市值", 0)) for p in positions)
    total_profit = total_mv - total_cost
    win_count = sum(1 for p in positions if (p.get("market_value", p.get("市值", 0)) > p.get("shares", p.get("份额", 0)) * p.get("avg_cost", p.get("成本", 0))))
    win_rate = (win_count / len(positions) * 100) if positions else 0

    summary_headers = ["指标", "数值"]
    for col, header in enumerate(summary_headers, start=1):
        cell = ws2.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = border

    summary_data = [
        ("总市值", f"¥{total_mv:,.2f}"),
        ("总成本", f"¥{total_cost:,.2f}"),
        ("总盈亏", f"¥{total_profit:,.2f}"),
        ("持仓数", len(positions)),
        ("胜率", f"{win_rate:.1f}%"),
    ]

    for row_idx, (label, value) in enumerate(summary_data, start=2):
        ws2.cell(row=row_idx, column=1, value=label).border = border
        ws2.cell(row=row_idx, column=2, value=value).border = border

    ws2.column_dimensions["A"].width = 12
    ws2.column_dimensions["B"].width = 20

    # ── Sheet 3: 因子评分 ────────────────────────────────────────────────────
    ws3 = wb.create_sheet("因子评分")

    factor_headers = ["代码", "名称", "估值", "成长", "质量", "动量", "情绪", "总分"]
    for col, header in enumerate(factor_headers, start=1):
        cell = ws3.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = border

    # 尝试从 factor_engine 获取评分
    factor_scores = []
    try:
        from factor_engine import score_positions as _score_fn
        _positions = positions
        try:
            from dashboard_views._shared import load_positions as _lp
            _positions = _lp() or positions
        except Exception:
            pass
        if _positions:
            factor_scores = _score_fn(_positions) or []
    except Exception:
        pass

    if factor_scores:
        for row_idx, r in enumerate(factor_scores, start=2):
            row_data = [
                r.get("ts_code", "—"),
                r.get("name", "—"),
                r.get("valuation_score", r.get("valuation", 0)),
                r.get("growth_score", r.get("growth", 0)),
                r.get("quality_score", r.get("quality", 0)),
                r.get("momentum_score", r.get("momentum", 0)),
                r.get("sentiment_score", r.get("sentiment", 0)),
                r.get("total_score", r.get("total", 0)),
            ]
            for col, value in enumerate(row_data, start=1):
                cell = ws3.cell(row=row_idx, column=col, value=value)
                cell.border = border
                if col >= 3:
                    cell.number_format = "0.00"
    else:
        # 无评分数据时写入占位
        ws3.cell(row=2, column=1, value="暂无因子评分数据").alignment = Alignment(horizontal="center")

    for col in ws3.columns:
        col_letter = get_column_letter(col[0].column)
        ws3.column_dimensions[col_letter].width = 12

    # 保存文件
    path = output_path or str(Path(__file__).parent.parent / "data" / "reports" / f"portfolio_{date.today().strftime('%Y%m%d')}.xlsx")
    Path(path).parent.mkdir(parents=True, exist_ok=True)
    wb.save(path)
    logger.info(f"Excel报告已生成: {path}")
    return path


if __name__ == "__main__":
    logging.basicConfig(level=logging.INFO)
    report = generate_weekly_report()
    path = save_report_as_md(report, "weekly")
    print(f"周报已生成: {path}")
    print(f"\n报告预览:\n{report[:500]}...")